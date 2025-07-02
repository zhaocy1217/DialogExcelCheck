import os
import re
import traceback
from collections import defaultdict
from datetime import timedelta, datetime
#from fcntl import F_SEAL_SHRINK
from openpyxl import load_workbook
from typing import Optional
from collections import Counter

import requests


# 记录aps表上各个词组需要过滤的id
cn_special_I_ids = {'refund_faq_answer4','refund_faq_answer3',100078, 100620, 100621, 100622, 121025, 121485, 170555, 208187, 211001, 211002, 211003, 211004, 211005, 211006, 211008, 211009, 211010, 211011, 211012, 211013, 211015, 211016, 211017, 211018, 211019, 211020, 211022, 211023, 211024, 211025, 211026, 211027, 211030, 211031, 211032, 211033, 211034, 211035, 211037, 211038, 211039, 211040, 211041, 211042, 211044, 211045, 211046, 211047, 211048, 211049, 211051, 211052, 211053, 211054, 211055, 211056, 211060, 211061, 211062, 211063, 211064, 211065, 211066, 211067, 211068, 211069, 211070, 211071, 211072, 211073, 211074, 211075, 211076, 211077, 211078, 211079, 211080, 211081, 211082, 211083, 211088, 211089, 211090, 211091, 211092, 211093, 211094, 211095, 211096, 211097, 211098, 211099, 211100, 211101, 211102, 211103, 211108, 211109, 211110, 211111, 211112, 211113, 211114, 211115, 211116, 211117, 211118, 211119, 211120, 211121, 211122, 211123, 211124, 211125, 211126, 211127, 211128, 211129, 211130, 211131, 211132, 211133, 211134, 211135, 211136, 211137, 211138, 211139, 211140, 211141, 211142, 211143, 211144, 211145, 211146, 211147, 211148, 211149, 211150, 211151, 211152, 211153, 211154, 211155, 211156, 211157, 211158, 211159, 211160, 211161, 211162, 211163, 211164, 211165, 211166, 211167, 211168, 211169, 211170, 211171, 211172, 211173, 211174, 211175, 211225, 211226, 211227, 211228, 211229, 211230, 211231, 211232, 211233, 211234, 211235, 211236, 211237, 211238, 211239, 320266, 320269, 320273, 320293, 320294, 320297, 320353, 370097, 370098, 370099, 372108, 400402, 400406, 430561, 450124, 450127, 455027, 455028, 455031, 455032, 455033, 455034, 455035, 455036, 456233, 456501, 456552, 458116, 458829, 458830, 458831, 309022, 309023, 309024, 309025, 309026, 309027, 309028, 309029, 309030, 309031, 309032, 309033, 800805, 803032, 803033, 803034, 803035, 'alliance_science_name001', 'alliance_science_name004', 'alliance_science_name006', 'alliance_science_name008', 'alliance_science_desc002', 'alliance_science_desc003', 'alliance_science_desc005', 'alliance_science_desc007', 'alliance_science_desc009', 'alliance_science_desc010', 'season_alliance_milestone_name003', 'season_alliance_milestone_name009', 'season_alliance_milestone_name013', 'season_alliance_milestone_name017', 'season_person_building_name001', 'season_person_building_name002', 'season_person_building_name003', 'season_person_building_name004', 'season_effect_name94001', 'season_effect_name94004', 'season_effect_name94006', 'season_effect_name94008', 'season_alliance_building_name40000', 'season_alliance_building_name40100', 'season_alliance_building_name40200', 'season_alliance_building_name40300', 'season_ui_desc049', 2000265, 2000266, 2000268, 2000269, 2000270, 2000272, 2000281, 2000284, 2000286, 2000288, 2000290, 2000295, 2000296, 2000299, 2000300, 2000301, 2000302, 2000303, 2000304, 2000305, 2000306, 2000307, 2000308, 2000309, 2000310, 2000311, 2000312, 2000313, 2000314, 2000315, 2000316, 2000370, 2000551, 2000810, 2000877, 2700007, 2700012, 2700013, 2700014, 'mail_content10082', 'snow_season_eventdesc1', 'activity_desc_99011', 'giftbag_name730501', 'giftbag_name730502', 'giftbag_name730504', 'giftbag_name730505', 'giftbag_name730507', 'giftbag_name730510', 'activity_desc_99026', 2900035, 2900036, 'vip_effect_desc_10001', 'item_name_541201', 'item_name_541302', 'item_name_541401', 'item_name_541402', 'item_name_541403', 'item_name_542201', 'item_name_542301', 'item_name_542401', 'item_name_542402', 'item_name_542403', 'item_name_543201', 'item_name_543301', 'item_name_543401', 'item_name_543402', 'item_name_543403', 'item_name_544201', 'item_name_544301', 'item_name_544401', 'item_name_544402', 'item_name_544403', 'delete_account_content_02', 'tech_duel_name_5', 'head_copy_id', 'uav_chip_name_1', 'uav_chip_name_2', 'uav_chip_name_3', 'uav_chip_name_4', 'uav_chip_name_5', 'uav_chip_name_9', 'uav_chip_name_10', 'uav_chip_name_11', 'uav_chip_name_12', 'uav_chip_name_13', 'uav_chip_name_17', 'uav_chip_name_18', 'uav_chip_name_19', 'uav_chip_name_20', 'uav_chip_name_21', 'uav_chip_name_25', 'uav_chip_name_26', 'uav_chip_name_27', 'uav_chip_name_28', 'uav_chip_name_29', 'dispatch_des014', 'vip_oldgift_desc1', 'tech_name_14_1', 'tech_name_14_2', 'tech_name_14_3', 'tech_name_14_4', 'tech_name_14_5', 'tech_name_14_6', 'tech_name_14_7', 'tech_name_14_8', 'tech_name_14_9', 'tech_name_14_10', 'tech_desc_14_11', 'tech_desc_14_12', 'tech_desc_14_13', 'tech_desc_14_14', 'tech_desc_14_15', 'tech_desc_14_16', 'tech_desc_14_17', 'tech_desc_14_18', 'tech_desc_14_19', 'tech_desc_14_20', 'tech_desc_15_11', 'tech_desc_15_12', 'tech_desc_15_13', 'tech_desc_15_14', 'tech_desc_15_15', 'tech_desc_15_16', 'tech_desc_15_17', 'tech_desc_15_18', 'tech_desc_15_19', 'tech_desc_15_20', 'tech_desc_16_11', 'tech_desc_16_12', 'tech_desc_16_13', 'tech_desc_16_14', 'tech_desc_16_15', 'tech_desc_16_16', 'tech_desc_16_17', 'tech_desc_16_18', 'tech_desc_16_19', 'tech_desc_16_20', 'tech_name_15_1', 'tech_name_15_2', 'tech_name_15_3', 'tech_name_15_4', 'tech_name_15_5', 'tech_name_15_6', 'tech_name_15_7', 'tech_name_15_8', 'tech_name_15_9', 'tech_name_15_10', 'tech_name_16_1', 'tech_name_16_2', 'tech_name_16_3', 'tech_name_16_4', 'tech_name_16_5', 'tech_name_16_6', 'tech_name_16_7', 'tech_name_16_8', 'tech_name_16_9', 'tech_name_16_10', 'overview_3', 'sevenday_event_des35', 'sevenday_event_des37', 'vip_s_connecttion_content', 'trialtower_yijian_08', 'vip_base_skin_rules1', 'vip_base_skin_rules2', 'vip_base_skin_rules14', 'vip_base_skin_rules20', 'vip_base_skin_rules27', 'vip_base_skin_rules33', 
'activity_tips_8001', 'season_s2_name', 'season_s2_mail_desc005', 'item_name520027', 'vip_expire_inform_title', 'vip_expire_inform', 'activity_desc_99087', 'giftbag_name_630018', 'season_s3_update_notice_13', 'season_s3_update_notice_14','item_desc612001', 'effectNumber_name_90029', 'vip_base_skin_desc1', 'vip_base_skin_desc2', 'vip_base_skin_desc11', 'hero_equip_skill_name_501940', 'hero_equip_skill_name_501950', 'item_name_611010', 'item_name_611010_1', 'item_name_611011', 'season_s2_end_trends_name02', 'activity_desc_99102', 'tech_name_12_01', 'tech_name_12_02', 'tech_name_12_03', 'item_name_611014', 'item_name_611015', 'item_name_611016', 'Desert_strom_tips1047', 'goldbrick_jp_desc', 'decoration_name16019', 'decoration_desc16019'}
cn_special_blank_ids = {100165, 100166, 310137, 'season_tips005', 'season_tiles_popui_info004', 'season_effect_name94018', 'new_city_activity_battle_tips1032', 'red_pocket_desc23', 'item_desc_620102', 'worker_hall_desc2 ', 'worker_hall_desc3', 'worker_hall_desc4', 'worker_hall_desc5', 'worker_hall_desc6', 'season_ice_survivor_10', 'soldier_death_rule_panel_3', 'giftbag_name_630018'}
cn_special_steel_food_ids = {'hero_equip_skill_name_500770','season_s4_activity1200010_story_desc','season_s4_activity_plot27','season_s4_plot_monster_906802', 'season_s4_monster_lang_desc', 100776, 180234, 210111, 211030, 211033, 211037, 211040, 211044, 211047, 211051, 211054, 211057, 220290, 221005, 250003, 251005, 260005, 450058, 450059, 450060, 450061, 450062, 450063, 450064, 450065, 450066, 450067, 500257, 800506, 800712, 801006, 801432, 801616, 'resource_name001', 'season_ui_desc048', 'season_trends_desc_101', 811037, 811070, 811131, 811149, 811158, 2000002, 2000004, 2000083, 2000096, 2000097, 2000098, 2000099, 2000131, 2000132, 2000182, 2000183, 2000184, 2000185, 2000200, 2000626, 2000649, 2000649, 2000779, 2010112, 2010112, 2500610, 2500710, 2501710, 'E100005', 'trialtower_004', 'trialtower_012', 'trade_person_tips1014', 'trade_person_tips1014', 'tech_desc_13_1', 'tech_desc_13_1', 'newparty_monster_crithit2', 'overview_effect_name_2027', 'hero_challenge_name5', 'hero_equip_skill_name_501760', 'hero_equip_skill_desc_501770', 'hero_equip_desc_501760', 'hero_equip_desc_501770', 'hero_equip_desc_501760_new', 'season_s2_food_activity_04', 'season_s2_food_activity_05', 'season_s2_food_activity_06', 'season_s2_food_activity_07', 'season_s2_food_activity_47', 'season_s2_food_activity_51', 'season_s2_activity_story_desc04', 'season_s2_food_activity_77', 'season_s2_food_activity_84', 'season_s2_food_activity_87', 'season_s2_food_activity_91', 'season_s2_food_activity_92', 'season_s2_food_activity_93', 'season_s2_food_activity_94', 'season_s2_food_activity_97', 'season_s2_food_activity_99', 'season_s2_food_activity_100', 'season_s2_food_activity_103', 'season_s2_activity_100036_plot_03', 'season_s2_activity_100036_plot_04', 'tech_special_desc_002', 'tech_special_desc_002', 'hero_equip_skill_name_500650', 'season_s2_activity_1000036_tips01', 'hero_brief_info_50007', 'tech_name_17_21', 'dominator_plot_31', 'dominator_cure_desc_7', 'item_name_800007'}
cn_special_castle_ids = {'season_s4_plot_905903', 'season_s4_city_desc07', 'season_ui_desc050', 'activity_99144_ABTIPS_3','season_activity_story_desc001', 2000467, 2000469, 2000812, 2000845, 'item_name610121', 'item_desc610121', 'item_desc610130', 'item_desc610135', 'item_desc610143', 'item_desc610147', 'item_name_611001', 'item_name_611006', 'item_desc_611001', 'item_desc_611006', 'activity_desc_99067', 'item_desc610142', 'decoration_skill_desc3', 'item_desc610151', 'item_desc610155', 'item_desc610159', 'item_desc612001', 'effectNumber_name_90029', 'item_desc610162', 'item_desc610164', 'item_desc610173', 'item_name_611014', 'item_desc_611014', 'activity_desc_99108', 'activity_thxgiv_castleLv1_desc', 'activity_thxgiv_castleLv2_desc', 'season_s3_activity1000067_story_desc', 'season_s3_hero_40006_story', 'item_desc610178', 'item_desc610179', 'decoration_skill_desc20', 'activity_desc_99115', 'item_name_611009'}
cn_special_above_below_ids= {'champion_duel_tips1057', 'season_s3_update_notice_9'}
cn_special_certificate_id={'season_s4_activity1200003_story_desc', 'season_s4_goods_name655028','decoration_name35004','season_s4_story_desc', 'item_name653113', 'dominator_story_hawk_content_6', 'snow_season_UI0008', 'activity_desc_99048', 'activity_desc_99012', 372273, 'season_s2_food_activity_38', 372287, 'item_name653006', 'wenjuan_02', 'activity_desc_99047', 'activity_desc_99061', 'snow_season_eventtitle4', 'season_mail_desc011', 458840, 'item_desc653117', 'activity_desc_99088', 'item_name653042', 'season_s3_rules_desc02', 'item_desc710055', 'activity_desc_99080', 'item_name710042', 'activity_slots_tips031', 'world_trends_event_S1_description', 'season_s2_food_activity_30', 803001, 'item_name653043', 'activity_sports_useitem_desc1', 'snow_season_eventtitle9', 'activity_newparty_desc25', 'season_yonghengzhiwang_duihua_2', 'activity_desc_99067', 'season_mail_message_01', 'item_desc710054', 'item_desc710039', 'item_name653030', 'activity_desc_99050', 'item_name653054', 'activity_sports_uitips_005', 'giftbag_name730504', 'giftbag_name730503', 'snow_season_eventtitle2', 'activity_desc_99027', 'season_s2_activity_story_desc12', 'slots_box_name1_1', 'season_s2_food_activity_70', 'season_s2_food_activity_43', 'season_s3_mail_desc80266', 'item_desc653042', 'wenjuan_01', 'activity_sports_uitips_033', 'item_name653068', 'item_name653000', 'activity_newparty_desc26', 'activity_slots_tips031_1', 'slots_box_name1', 'activity_desc_99051', 'season_city_zhanling_2_2', 'season_s2_food_activity_83', 'music_slots_box_name1', 'item_name653062', 'giftbag_name730501', 'giftbag_name730506', 'giftbag_name730512', 'activity_desc_99026', 'activity_desc_99040', 'item_desc653042_huishou', 'season_s2_food_activity_78', 'giftbag_name730507', 'season3_comic_6', 'giftbag_name730510', 'wenjuan_mail_title_01', 'activity_desc_99081', 'snow_season_UI0014', 'item_desc710053', 'giftbag_name730509', 'activity_desc_99020', 'season_city_zhanling_7_4', 'world_trends_event_S2_description', 'season_s3_story_desc'}
cn_special_decorate_id={'item_name705501', 'decoration_colorful_skin_tips3', 'decoration_item_name3', 'activity_ValentineBuffDesc5', 'item_desc_706901', 'decoration_item_name2', 'item_desc_707601', 'item_desc705701', 'item_desc_706101', 'decoration_recruit_desc30', 'decoration_recruit_desc7', 'activity_ValentineBuffDesc3', 'item_name653102', 'item_desc_706801', 'decoration_colorful_skin_tips1', 'decoration_recruit_desc33', 'item_name653086', 'item_desc_707501', 'decoration_recruit_desc46', 'giftbag_name743596', 'giftbag_name743575', 'decoration_recruit_desc32', 'item_desc705011', 'giftbag_name743521', 'item_desc707002', 'item_desc704011', 'season_s2_activity1000054_des', 'decoration_recruit_desc5', 'item_desc705201', 'item_desc705051', 'decoration_recruit_desc38', 'item_desc_707101', 'activity_slots_tips036', 'item_name653087', 'giftbag_name733521', 'decoration_recruit_desc21', 'season_s3_adventurer_dialog05', 'item_desc_706501', 'item_desc707001', 'decoration_recruit_desc47', 'decoration_recruit_desc14', 'giftbag_name743584', 'item_desc705501', 'item_desc706021', 'activity_desc_99137', 'giftbag_name743593', 'item_desc_707401', 'decoration_recruit_desc41', 'item_desc_706601', 'giftbag_name743565', 'building_desc_new103511000', 'item_desc_706301', 'item_name705201', 'decoration_recruit_desc22', 'decoration_item_desc1', 'decoration_colorful_skin_tips2', 'item_name705301', 'building_desc_new103513000', 'item_desc703031', 'item_desc705801', 'item_desc704031', 'item_desc_706201', 'item_desc703041', 'activity_ValentineBuffDesc1', 'decoration_colorful_skin_tips4', 'item_desc705061', 'item_desc705031', 'activity_ValentineBuffDesc4', 'season_s2_callback_tips_9', 'deco_expire_inform', 'giftbag_name743541', 'item_name653084', 'item_desc705041', 'item_name653085', 'item_desc703021', 'decoration_item_name1', 'giftbag_desc743521', 'item_desc705901', 'giftbag_name743587', 'item_desc704041', 'item_desc705071', 'season_s2_callback_tips_9_new', 'item_desc704021', 'item_desc_707201', 'item_desc705021', 'giftbag_name743570', 'item_name653101', 'item_desc703011', 'season_s1_activity1000055_des', 'decoration_item_desc2', 'item_desc_706401', 'activity_ValentineBuffDesc2', 'giftbag_name743580', 'decoration_item_name4', 'item_desc710076', 'deco_expire_inform_title', 'item_desc705081', 'item_desc_707301', 'item_desc707004', 'giftbag_name743545', 'season_s2_callback_tips_8', 'giftbag_name743590', 'activity_desc_99115'}
cn_special_base_effect_skin_ids = {'item_desc610184'}   
cn_special_team_ids = {'season_s4_activity1200012_story_desc', 221056, 'battlesystem_info1_new', 1500930, 'city_trade_tips1017', 1501830, 210185, 210186, 210187, 1501330, 'skill_desc_402020_new', 'report_prop25', 2000539, 1501730, 'science_221043_new', 1401130, 1501230, 457009, 1400630, 1500730, 1501630, 372287, 'science_221055_new', 220358, 1401030, 1501130, 'city_trade_tips1012', 'battlesystem_info1', 'power_display_new_002', 801105, 457554, 457555, 457556, 457557, 457558, 457559, 457560, 'science_221044_new', 457561, 457563, 1401430, 1500630, 141150, 1501530, 'skill_desc_402020', 'report_prop24', 1401830, 1501030, 'science_221054_new', 221043, 221044, 221045, 221046, 221047, 1501430, 221049, 'report_helper_strategy_34_detail', 221051, 221053, 221054, 221055}
cn_special_min_ids = {'season_s3_update_notice_14'}



class LocalizeChecker():
    excel_name = "APS_Dialog.xlsm"
    def __init__(self):
        self.name_id = NoticeManager.name_id
        # 记录aps表的各个字段，检查字段删除（新增字段时需更新）
        self.aps_fields = ['id', '1-不需要翻译', 'CN', 'gn_CN', 'CN_patch', '预期时间', '优先级', '签名', '说明1', '说明2', '词条备注：通配符例句、说明', 'crowdin格式专用（勿动）']

    async def check(self, local_path, funcs, is_warn: bool=False, is_pub: bool = False):
        try:
            self.error_message = ""
            self.error_brief = list()
            self.path = local_path

            # 确保 funcs 是一个可迭代对象
            if not isinstance(funcs, (list, tuple)):
                funcs = [funcs]
            
            # 依次执行每个函数
            for func in funcs:
                if func == self.check_APS_Dialog_CN:
                    await func(is_pub)
                else:
                    await func()
        except:
            self.error_message += traceback.format_exc()
            # 填写需要默认通知用户的飞书id
            self.error_usrs = [self.name_id.get('赵超跃')]

        finally: 
            return
            error_content = "**错误简要**：\n" + '\n'.join(self.error_brief)
            if self.error_message:
                if not is_warn:
                    # 发送检查错误, 本地化错误需要专门发送到本地化群
                    self.send_error(error_content=error_content, url="https://open.feishu.cn/open-apis/bot/v2/hook/e3069555-89da-4679-89ee-3bb5ab7bb1f2")# url填写飞书机器人的url

                else:
                    # 发送检查警告
                    self.send_warning(error_content=error_content, url="https://open.feishu.cn/open-apis/bot/v2/hook/e3069555-89da-4679-89ee-3bb5ab7bb1f2")
            return self.error_message

    

    async def check_CN(self, local_path, is_pub: bool = False):
        """
        local_path: 本地存储配置的路径
        is_pub: 是否是发版检查
        """
        self.error_usrs = [self.name_id.get('赵超跃')]#[self.name_id.get('阳振岳'), self.name_id.get('苏湘鹏'), self.name_id.get('赵超跃')] # 填写需要通知用户的飞书id
        return await self.check(
            funcs=self.check_APS_Dialog_CN,
            local_path=local_path,
            is_pub=is_pub
        )

    async def check_APS_Dialog_CN(self, is_pub: bool = False):
        # 检查条件在这更新
        error_checks = [
            # 检查是否有大写I,过滤VIP，ID
            {
                "condition": lambda cn, id: 'I' in cn.replace('VIP', '').replace('ID', '') and id not in cn_special_I_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】存在I\n"
            },
            # 检查开头结尾是否有空格
            {
                "condition": lambda cn, id: not re.match(r"^(?! )(?:(?! $).)*$", cn, re.DOTALL) and id not in cn_special_blank_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】开头或结尾存在空格\n"
            },
            # 检查是否存在手动换行
            {
                "condition": lambda cn, id: re.search(r"(?<!\\)\n", cn),
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】中存在换行符\n"
            },
            # 检查出现钢铁
            {
                "condition": lambda cn, id: "钢铁" in cn and id not in cn_special_steel_food_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现钢铁\n"
            },
            # 检查出现食物
            {
                "condition": lambda cn, id: "食物" in cn and id not in cn_special_steel_food_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现食物\n"
            },
            # 检查出现城堡
            {
                "condition": lambda cn, id: ("城堡" in cn or "主城" in cn  or "城内" in cn) and id not in cn_special_castle_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】基地名称不正确\n"
            },
            # 检查出现数字\d{2,}以上
            {
                "condition": lambda cn, id: re.search(r"\d{2,}以上", cn) and not re.search(r"\d{2,}及以上", cn) and id not in cn_special_above_below_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现数字以上\n"
            },
            # 检查出现数字\d{2,}以下
            {
                "condition": lambda cn, id: re.search(r"\d{2,}以下", cn) and not re.search(r"\d{2,}及以下", cn) and id not in cn_special_above_below_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现数字以下\n"
            },
            # 检查出现诺亚
            {
                "condition": lambda cn, id: "诺亚" in cn,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现诺亚\n"
            },
            # 检查时间是否符合格式要求
            {
                "condition": lambda cn, id, expect_time: not isinstance(expect_time, datetime),
                "error_message": lambda id, expect_time: f"【{excel}】【id={id}】【预期时间={expect_time}】不符合时间格式\n"
            },
            # 检查出现劵或卷
            {
                "condition": lambda cn, id: ("劵" in cn or "卷" in cn) and id not in cn_special_certificate_id,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现劵或卷\n"
            },
            # 检查只能出现装饰物，不能只出现装饰
            {
                "condition": lambda cn, id: "装饰" in cn.replace("装饰物", "") and id not in cn_special_decorate_id,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】使用装饰而不是装饰物\n"
            },
            # 检查不能出现总部装扮、总部皮肤、总部特效、特效皮肤
            {
                "condition": lambda cn, id: any(term in cn for term in ["总部装扮", "总部皮肤", "总部特效", "特效皮肤"]) and id not in cn_special_base_effect_skin_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现总部装扮或总部皮肤或总部特效或特效皮肤\n"
            },
            # 检查不能出现车队
            {
                "condition": lambda cn, id: "车队" in cn and id not in cn_special_team_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现车队\n"
            },
            # 检查不能出现Min\min
            {
                "condition": lambda cn, id: ("min" in cn or "Min" in cn) and id not in cn_special_min_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现Min\\min\n"
            },
             # 检查不能出现大地图、野外
            {
                "condition": lambda cn, id: "大地图" in cn or "野外" in cn,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现大地图、野外\n"
            },
             # 检查不能出现资源点
            {
                "condition": lambda cn, id: "资源点" in cn,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现资源点\n"
            },
             # 检查不能出现雷达事件
            {
                "condition": lambda cn, id: "雷达事件" in cn,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现雷达事件\n"
            },
             # 检查不能出现士气值
            {
                "condition": lambda cn, id: "士气值" in cn,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现士气值\n"
            },
            # 检查不能出现本服、跨服
            {
                "condition": lambda cn, id: "本服、跨服" in cn,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】出现本服、跨服\n"
            },
        ]

        # 发版检查条件在这更新
        publish_error_checks = [
            # 检查开头结尾是否有空格
            {
                "condition": lambda cn, id: not re.match(r"^(?! )(?:(?! $).)*$", cn, re.DOTALL) and id not in cn_special_blank_ids,
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】开头或结尾存在空格\n"
            },
            # 检查是否存在手动换行
            {
                "condition": lambda cn, id: re.search(r"(?<!\\)\n", cn),
                "error_message": lambda id, cn: f"【{excel}】【id={id}】【cn={cn}】中存在换行符\n"
            },
            # 检查时间是否符合格式要求
            {
                "condition": lambda cn, id, expect_time: not isinstance(expect_time, datetime),
                "error_message": lambda id, expect_time: f"【{excel}】【id={id}】【预期时间={expect_time}】不符合时间格式\n"
            },
        ]

        excel = self.excel_name if hasattr(self, 'excelName') else "APS_Dialog.xlsm" #"APS_Dialog.xlsm"
        sheet = "CN"
        ids = defaultdict(list)
        file_path=os.path.join(self.path, excel)
        column_order, data_list = get_sheet_column_data(
            path=file_path,
            sheet_name=sheet,
            field_row_index=0
        )
        current_fields = list(column_order.keys())
        counter1 = Counter(current_fields)
        counter2 = Counter(self.aps_fields)
        # 找出 list1 中比 list2 缺少的元素
        missing_list = counter2 - counter1  # list2 中多余的元素
        if missing_list:
            missing_list = list(dict(missing_list).keys())
            self.add_error(
                error_message = f'【{excel}】缺少【{missing_list}】列\n'
            )
            return

        id_index = 0
        cn_index = column_order.get("CN")
        sign_index = column_order.get("签名")
        expect_time_index = column_order.get("预期时间")
        if is_pub:
            error_checks = publish_error_checks
        for row in data_list[1:]:
            if id := row[id_index]:
                sign = row[sign_index]
                ids[id].append(sign)
                cn = str(row[cn_index])
                expect_time = row[expect_time_index]
                for check in error_checks:
                    # 动态检查需要传递的参数数量
                    condition_args = check["condition"].__code__.co_varnames#获取函数中所有局部变量的名称
                    if len(condition_args) == 2:  # 只需要 cn 和 id
                        if check["condition"](cn, id):
                            cn_special_castle_ids.add(id)
                            self.add_error(
                                error_message=check["error_message"](id, cn),
                                sign=sign
                            )
                    elif len(condition_args) == 3:  # 需要 cn, id 和 expect_time
                        if check["condition"](cn, id, expect_time):
                            self.add_error(
                                error_message=check["error_message"](id, expect_time),
                                sign=sign
                            )

        for id, sign in ids.items():
            if len(sign) > 1:
                error_message = f'APS_Dialog的id重复检查: 【id ={id}】重复'
                self.add_error(
                    error_message=error_message,
                    # 通知最后一个
                    sign=sign[-1]
                )

    def add_error(self, error_message:str, sign:str=""):
        self.error_brief.append(error_message)
        self.error_message += error_message
        sign_id = self.name_id.get(sign)
        md = self.name_id.get('田明东')
        if sign and sign_id and sign_id not in self.error_usrs:
            self.error_usrs.append(sign_id)
        if md not in self.error_usrs:
            self.error_usrs.append(md)

    async def warn_CN(self, local_path):
        self.error_usrs = [self.name_id.get('田明东'), self.name_id.get('阳振岳'), self.name_id.get('苏湘鹏')]
        return await self.check(
            funcs=self.warn_APS_Dialog_CN,
            local_path=local_path,
            is_warn=True
        )

    async def warn_APS_Dialog_CN(self):
        excel = self.excel_name if hasattr(self, 'excelName') else "APS_Dialog.xlsm" #"APS_Dialog.xlsm"
        sheet = "CN"
        file_path=os.path.join(self.path, excel)
        column_order, data_list = get_sheet_column_data(
            path=file_path,
            sheet_name=sheet,
            field_row_index=0
        )
        id_index = 0
        need_tarns_index = 1
        sign_index = column_order.get("签名")
        expect_time_index = column_order.get("预期时间")
        today = datetime.today()
        today = datetime(today.year, today.month, today.day)
        for row in data_list[1:]:
            if id := row[id_index]:
                sign = row[sign_index]
                need_tarns = row[need_tarns_index]
                # 在预期时间的前一天检查
                expect_time = row[expect_time_index]
                expect_date = datetime(expect_time.year, expect_time.month, expect_time.day)
                if today + timedelta(days=1) == expect_date:
                    if need_tarns != None:
                        need_tarns = int(need_tarns)
                    if need_tarns == 1:
                        error_msg = f"【{excel}】【id={id}】【expect_time={expect_time}】【1-不需要翻译={need_tarns}】有词条没有改成0\n"
                        self.add_error(
                                error_message=error_msg,
                                sign=sign
                            )
    
    def send_error(self, error_content: str, url:str):
        # 发送错误报告到通知群或指定群
        NoticeManager().send_file_notice(
            url=url,
            title="错误通知",
            content=error_content, 
            is_error=True,
            error_usrs={self.name_id.get('赵超跃'), self.name_id.get('苏湘鹏')}# 填写需要通知用户的飞书id
        )
    
    def send_warning(self, error_content: str, url:str):
        # 发送警告报告到通知群
        NoticeManager().send_file_notice(
            url=url,
            title="警告通知",
            content=error_content,  
            is_error=False,
            error_usrs={self.name_id.get('赵超跃'), self.name_id.get('苏湘鹏')}# 填写需要通知用户的飞书id
        )
    
class NoticeManager():
    method = 'post'
    headers = {
        'Content-Type': 'application/json'
    }
    usr_id = {
        "192.168.50.174": "7383886191577563138",
        "192.168.50.136": "7356053525241266204",
        "192.168.82.188": "7363481092312317955",
        "192.168.50.131": "7233978831617572865",
        "192.168.60.115": "7356053525241266204",
        "192.168.50.15": "7356053525241266204",
        "192.168.50.136": "7356053525241266204",
        "255.255.255.255": "7330847096047616003",
        "192.168.50.114": "7347518430268555268",
        "192.168.60.124": "7347518430268555268",
        "192.168.83.33": "7363481092312317955",
        "192.168.50.248": "7389901589062320156",
        "192.168.50.70": "7330847096047616003",
        "192.168.83.66": "7330847096047616003",  
    }

    name_id = {
        "赵超跃": "7480015286726115347",
        "阳振岳": "7389901589062320156",
        "苏湘鹏": "7319691619463102467",
        "张瑞": "7407267753591586819",
        "杨凯允": "7418032801147142147",
        '谭广能': '7022857211525234691',
        '王秋晖': '7281589147767275521',
        '田明东': '7280716964732829699',
        'Jason': '7225427920553213980',
        'jason': '7225427920553213980',
        '陈禹同': '7341213139674955778',
        '禹同': '7341213139674955778',
        '汪昌海': '7126718782088888324',
        '昌海': '7126718782088888324',
        '唐小凡': '7168273299829653507',
        '刘泽宇 ': '7280350202438189057',
        '苏浩波': '7311162129275944964',
        '刘经纬': '7294078862072266755',
        '朱兵': '7168273200103424028',
        '代可君': '7218757583815852060',
        '麻琳': '7218757589222686722',
        '佳豪': '7292967315640188929',
        '龚佳豪': '7292967315640188929',
        'Jack': '6900534029628473346',
        'jack': '6900534029628473346',
        '心悦': '7295576583639826435',
        '胡心悦': '7295576583639826435',
        '李卓群': '7325403228282060803',
        '王诗杨': '7218757580565364738',
        '叶枫': '7369042612079312899',
        '梁霄': '7168274928364519425',
        '王春晓': '7389082429509910530',
        '黎俊铭': '7404674799434072068',
        '陈亚珣': '7412466118919553027',
        '李帆': "7330847096047616003",
        '李进武': "7356053525241266204",
        '刘航': "7363481092312317955",
        '姚逸宁': "7457460885003780098",
        '陈嘉伟': "7457460869476417555",
        '马瑶': "7202804136942780418",
        '张俸铭': "7215790035071729668",
        "杨超": "6978670367170691074",
        "蔡英杰": "7313756880567058435",
        "李永杰": "7454024529842716675",
        "王阳明": "7459966893919846404",
        '杨佳蔚': '7131939406029570050',
        '赵聪慧': '6978671554364178436',
        '王慧茹': '6978670341132386306',
        '谢子钰': '7371274361744080900',
        '郭小明': '7002055487910871044',
        '王子文': '7202432063675760668',
        '刘强': '6972502602839801857',
        '魏泽行': '7124145008772513796',
        '林子禹': '6978674875120517122',
        '苏卫东': '7049555408477749249',
        '黄鼎': '7360519197648453636',
        '王鑫滈': '7106704851811319809',
        '史苏明': '7282952866506260481',
        '苏赫': '6978679225003819010',
        '李小松': '6989436802142650396',
        '郭阳': '6994999679998033948',
        '雷浩然': '6978679394969845762',
        "孙金成": '7392420567398924316',
        '杨凯允': '7418032801147142147',
        '王律': '7328001261696253955',
        '王乐': '7383886191577563138',
        '魏栋梁': '7347518430268555268',
        "张玉强": '7343439571297894401',
        "高运": '7324150993162092548',
        "张建军": '7317842336485326851',
        "王振华": '7455141192876326915',
        "刘继丰": '7480746472863252483',
        "付项哲": '7490392722126143492',
        "黄佳承": '7437336797694427164',
        "罗希彦": "7451432165647024132",
        "樊皓云": "7462199777667956738",
        "刘博语": "7482608283694481411",
        "胡哲": "7459231689765191699",
        "倪千惠": "7501527646409719811",
    }

    def send_file_notice(self, url:str, title: str, content: str = "", is_error: bool = False, error_usrs: set = None):
        """
        发送飞书通知
        url: 飞书通知的url
        title: 通知的标题
        content: 通知的内容
        is_error: 是否是错误通知
        error_usrs: 需要通知的用户
        """
        if error_usrs:
            error_at = self.get_error_at(error_usrs=error_usrs)
        else:
            error_at = self.get_error_at(error_usrs={self.name_id.get('阳振岳')})

        if is_error:
            template = "red" 
        else:
            template = "yellow"

        json = {
            "msg_type": "interactive",
            "card": {
                "elements": [{
                        "tag": "div",
                        "text": {
                                "content": content + error_at,
                                "tag": "lark_md"
                        }
                }],
                "header": {
                        "title": {
                                "content": title,
                                "tag": "plain_text"
                        },
                        "template": template
                }
            }
        }

        try:
            res = requests.request(method=self.method, url=url, headers=self.headers, json=json)
            print(res.text)
        except:
            traceback.format_exc()
    
    def get_error_at(self, error_usrs: set) -> str:
        """将需要通知的名单转化成配置中的飞书id"""
        error_at = ""
        if error_usrs:
            new_error_usrs = set()
            for error_usr in error_usrs:
                if name_id := self.name_id.get(error_usr):
                    new_error_usrs.add(name_id)
                
                else:
                    new_error_usrs.add(error_usr)

            for new_error_usr in new_error_usrs:
                error_at += f"<at id={new_error_usr}></at>"

        return error_at
    
def get_sheet_column_data(path: str, sheet_name: str, field_row_index: Optional[int]=None):
    """
    获取指定表的指定sheet的列索引和数据
    path: 文件路径
    sheet_name: 表名
    field_row_index: 字段行索引
    """
    workbook = load_workbook(path, data_only=True)
    column_order = dict()
    data_list=list()
    sheet = workbook[sheet_name]

    if field_row_index == None:
        field_row_index = 8

    for index, column in enumerate(sheet.iter_cols(values_only=True)):
        if field:= column[field_row_index]:
            column_order[field] = index

    for row in sheet.iter_rows(values_only=True):
        data_list.append(list(row))

    return column_order, data_list